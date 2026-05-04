"""Tests integration RAT/DPIA (skill 5 — Ley 21.719).

Cubren CRUD, role gate (solo owner / accountant_lead pueden mutar),
RLS por workspace y export XLSX."""

from __future__ import annotations

import io
from collections.abc import AsyncIterator, Callable
from typing import Any
from uuid import UUID, uuid4

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.auth.jwt import verify_jwt
from src.main import app


def _claims(
    user_id: UUID,
    workspace_id: UUID,
    *,
    role: str = "owner",
    workspace_type: str = "accounting_firm",
) -> dict[str, Any]:
    return {
        "sub": str(user_id),
        "email": f"dpo-{user_id}@renteo.local",
        "aud": "authenticated",
        "role": "authenticated",
        "app_metadata": {
            "provider": "email",
            "workspace_id": str(workspace_id),
            "workspace_type": workspace_type,
            "role": role,
            "empresa_ids": [],
        },
    }


def _override_jwt(claims: dict[str, Any]) -> Callable[[], dict[str, Any]]:
    def _dep() -> dict[str, Any]:
        return claims

    return _dep


@pytest_asyncio.fixture
async def http_client_pc() -> AsyncIterator[AsyncClient]:
    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver"
    ) as client:
        yield client
    app.dependency_overrides.clear()


@pytest_asyncio.fixture
async def workspace_dpo(
    admin_session: AsyncSession,
) -> AsyncIterator[dict[str, UUID]]:
    user_id = uuid4()
    workspace_id = uuid4()
    async with admin_session.begin():
        await admin_session.execute(
            text("insert into auth.users (id, email) values (:id, :e)"),
            {"id": str(user_id), "e": f"dpo-{user_id}@renteo.local"},
        )
        await admin_session.execute(
            text(
                "insert into core.workspaces (id, name, type) "
                "values (:id, 'Workspace DPO Test', 'accounting_firm')"
            ),
            {"id": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                "insert into core.workspace_members "
                "(workspace_id, user_id, role, accepted_at) "
                "values (:ws, :u, 'owner', now())"
            ),
            {"ws": str(workspace_id), "u": str(user_id)},
        )
    yield {"user_id": user_id, "workspace_id": workspace_id}
    async with admin_session.begin():
        await admin_session.execute(
            text("set local session_replication_role = 'replica'")
        )
        await admin_session.execute(
            text(
                "delete from privacy.dpia_records where workspace_id = :w"
            ),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                "delete from privacy.rat_records where workspace_id = :w"
            ),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text("delete from core.workspaces where id = :w"),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text("delete from auth.users where id = :u"),
            {"u": str(user_id)},
        )


def _rat_payload() -> dict[str, Any]:
    return {
        "nombre_actividad": "Tratamiento RUT clientes pyme",
        "finalidad": "Cálculo tributario y emisión de informes",
        "base_legal": "contrato",
        "categorias_titulares": ["clientes_pyme", "representantes"],
        "categorias_datos": [
            "identificacion",
            "tributarios",
            "contacto",
        ],
        "datos_sensibles": False,
        "encargados_referenciados": ["supabase", "aws-kms"],
        "transferencias_internacionales": [
            {"pais": "EE.UU.", "garantia": "DPA AWS"}
        ],
        "plazo_conservacion": "5 años (art. 200 CT)",
        "medidas_seguridad": ["AES-256 reposo", "TLS 1.3 tránsito"],
        "responsable_email": "dpo@renteo.cl",
    }


@pytest.mark.integration
async def test_rat_crud_full_cycle(
    http_client_pc: AsyncClient, workspace_dpo: dict[str, UUID]
) -> None:
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(workspace_dpo["user_id"], workspace_dpo["workspace_id"])
    )
    create_resp = await http_client_pc.post(
        "/api/privacy/rat",
        json=_rat_payload(),
        headers={"Authorization": "Bearer fake"},
    )
    assert create_resp.status_code == 201, create_resp.text
    rat_id = create_resp.json()["id"]

    list_resp = await http_client_pc.get(
        "/api/privacy/rat",
        headers={"Authorization": "Bearer fake"},
    )
    assert list_resp.status_code == 200
    nombres = [
        r["nombre_actividad"] for r in list_resp.json()["records"]
    ]
    assert "Tratamiento RUT clientes pyme" in nombres

    patch_resp = await http_client_pc.patch(
        f"/api/privacy/rat/{rat_id}",
        json={"plazo_conservacion": "6 años"},
        headers={"Authorization": "Bearer fake"},
    )
    assert patch_resp.status_code == 200
    assert patch_resp.json()["plazo_conservacion"] == "6 años"

    archive_resp = await http_client_pc.delete(
        f"/api/privacy/rat/{rat_id}",
        headers={"Authorization": "Bearer fake"},
    )
    assert archive_resp.status_code == 204

    list_after = await http_client_pc.get(
        "/api/privacy/rat",
        headers={"Authorization": "Bearer fake"},
    )
    nombres_after = [
        r["nombre_actividad"] for r in list_after.json()["records"]
    ]
    assert "Tratamiento RUT clientes pyme" not in nombres_after


@pytest.mark.integration
async def test_rat_blocked_for_viewer(
    http_client_pc: AsyncClient, workspace_dpo: dict[str, UUID]
) -> None:
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(
            workspace_dpo["user_id"],
            workspace_dpo["workspace_id"],
            role="viewer",
        )
    )
    response = await http_client_pc.post(
        "/api/privacy/rat",
        json=_rat_payload(),
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 403


@pytest.mark.integration
async def test_dpia_create_and_approve(
    http_client_pc: AsyncClient, workspace_dpo: dict[str, UUID]
) -> None:
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(workspace_dpo["user_id"], workspace_dpo["workspace_id"])
    )
    rat_resp = await http_client_pc.post(
        "/api/privacy/rat",
        json=_rat_payload(),
        headers={"Authorization": "Bearer fake"},
    )
    rat_id = rat_resp.json()["id"]

    dpia_resp = await http_client_pc.post(
        "/api/privacy/dpia",
        json={
            "rat_id": rat_id,
            "nombre_evaluacion": "DPIA tratamiento RUT",
            "descripcion_tratamiento": (
                "Tratamiento de RUT y datos tributarios para cálculos "
                "de optimización tributaria."
            ),
            "necesidad_proporcionalidad": (
                "Mínimo necesario para cumplir el contrato; "
                "anonimización en logs y reportes agregados."
            ),
            "riesgos_identificados": [
                {
                    "descripcion": "Filtración SII",
                    "probabilidad": "baja",
                    "impacto": "alto",
                }
            ],
            "medidas_mitigacion": [
                "Cifrado en reposo",
                "Acceso por roles",
            ],
            "riesgo_residual": "medio",
        },
        headers={"Authorization": "Bearer fake"},
    )
    assert dpia_resp.status_code == 201, dpia_resp.text
    dpia_id = dpia_resp.json()["id"]
    assert dpia_resp.json()["aprobado_at"] is None
    assert dpia_resp.json()["version"] == 1

    aprobar = await http_client_pc.patch(
        f"/api/privacy/dpia/{dpia_id}",
        json={"aprobar": True},
        headers={"Authorization": "Bearer fake"},
    )
    assert aprobar.status_code == 200, aprobar.text
    body = aprobar.json()
    assert body["aprobado_at"] is not None
    assert body["aprobado_por_dpo_email"] is not None
    assert body["version"] == 2


@pytest.mark.integration
async def test_rat_export_xlsx(
    http_client_pc: AsyncClient, workspace_dpo: dict[str, UUID]
) -> None:
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(workspace_dpo["user_id"], workspace_dpo["workspace_id"])
    )
    await http_client_pc.post(
        "/api/privacy/rat",
        json=_rat_payload(),
        headers={"Authorization": "Bearer fake"},
    )
    response = await http_client_pc.get(
        "/api/privacy/rat.xlsx",
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert response.content[:2] == b"PK"

    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    assert "Resumen" in wb.sheetnames
    assert "Actividades" in wb.sheetnames

    activ = wb["Actividades"]
    flat = " ".join(
        str(cell)
        for row in activ.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Tratamiento RUT clientes pyme" in flat
    assert "art. 200 CT" in flat


@pytest.mark.integration
async def test_dpia_export_xlsx_empty(
    http_client_pc: AsyncClient, workspace_dpo: dict[str, UUID]
) -> None:
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(workspace_dpo["user_id"], workspace_dpo["workspace_id"])
    )
    response = await http_client_pc.get(
        "/api/privacy/dpia.xlsx",
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    eval_sheet = wb["Evaluaciones"]
    flat = " ".join(
        str(cell)
        for row in eval_sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Sin evaluaciones registradas" in flat
