"""Tests integration del papel de trabajo cliente B (skill 9 — closure).

Verifican: descarga retorna XLSX válido con los datos de la empresa,
hash, fundamento legal, escenarios y SII sync; 404 cuando la empresa
no es accesible; audit log entry persistida."""

from __future__ import annotations

import io
import json
from collections.abc import AsyncIterator, Callable
from typing import Any
from uuid import UUID, uuid4

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.auth.jwt import verify_jwt
from src.main import app


def _claims(
    user_id: UUID,
    workspace_id: UUID,
    *,
    role: str = "accountant_lead",
    workspace_type: str = "accounting_firm",
) -> dict[str, Any]:
    return {
        "sub": str(user_id),
        "email": f"contador-{user_id}@renteo.local",
        "aud": "authenticated",
        "role": "authenticated",
        "app_metadata": {
            "provider": "email",
            "workspace_id": str(workspace_id),
            "workspace_type": workspace_type,
            "role": role,
            "empresa_ids": [],
        },
    }


def _override_jwt(claims: dict[str, Any]) -> Callable[[], dict[str, Any]]:
    def _dep() -> dict[str, Any]:
        return claims

    return _dep


@pytest_asyncio.fixture
async def http_client_papel() -> AsyncIterator[AsyncClient]:
    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver"
    ) as client:
        yield client
    app.dependency_overrides.clear()


@pytest_asyncio.fixture
async def empresa_con_diagnostico(
    admin_session: AsyncSession,
) -> AsyncIterator[dict[str, UUID]]:
    user_id = uuid4()
    workspace_id = uuid4()
    empresa_id = uuid4()
    rec_id = uuid4()
    scen_id = uuid4()
    async with admin_session.begin():
        await admin_session.execute(
            text("insert into auth.users (id, email) values (:id, :e)"),
            {"id": str(user_id), "e": f"papel-{user_id}@renteo.local"},
        )
        await admin_session.execute(
            text(
                "insert into core.workspaces (id, name, type) "
                "values (:id, 'Estudio Contable Demo', 'accounting_firm')"
            ),
            {"id": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                "insert into core.workspace_members "
                "(workspace_id, user_id, role, accepted_at) "
                "values (:ws, :u, 'accountant_lead', now())"
            ),
            {"ws": str(workspace_id), "u": str(user_id)},
        )
        await admin_session.execute(
            text(
                """
                insert into core.empresas
                    (id, workspace_id, rut, razon_social, giro,
                     regimen_actual)
                values (:e, :ws, '11111111-1', 'Empresa Papel SpA',
                        'Comercio al por menor', '14_d_3')
                """
            ),
            {"e": str(empresa_id), "ws": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                """
                insert into core.recomendaciones
                    (id, workspace_id, empresa_id, tax_year, tipo,
                     descripcion, fundamento_legal, ahorro_estimado_clp,
                     disclaimer_version, engine_version,
                     inputs_snapshot, outputs,
                     rule_set_snapshot, tax_year_params_snapshot,
                     rules_snapshot_hash, created_by)
                values
                    (:id, :ws, :emp, 2026, 'cambio_regimen',
                     'Cambio a 14 D N°3 con plan retiros 30%',
                     cast(:fund as jsonb), 1500000,
                     'disclaimer-recomendacion-v1', 'engine-vTEST',
                     cast(:inp as jsonb), cast(:out as jsonb),
                     cast('{}' as jsonb), cast('{}' as jsonb),
                     'a' || repeat('1', 63), :uid)
                """
            ),
            {
                "id": str(rec_id),
                "ws": str(workspace_id),
                "emp": str(empresa_id),
                "uid": str(user_id),
                "fund": json.dumps(
                    [
                        {"texto": "art. 14 D N°3 LIR"},
                        {"texto": "Circular SII 53/2025"},
                    ]
                ),
                "inp": json.dumps(
                    {"regimen_actual": "14_a", "tax_year": 2026}
                ),
                "out": json.dumps(
                    {
                        "veredicto": {
                            "regimen_actual": "14_a",
                            "regimen_recomendado": "14_d_3",
                        }
                    }
                ),
            },
        )
        await admin_session.execute(
            text(
                """
                insert into core.escenarios_simulacion
                    (id, workspace_id, empresa_id, tax_year, nombre,
                     regimen, inputs, outputs,
                     engine_version, created_by,
                     rule_set_snapshot, tax_year_params_snapshot,
                     rules_snapshot_hash)
                values
                    (:id, :ws, :emp, 2026, 'Plan A',
                     '14_d_3', cast(:inp as jsonb), cast(:out as jsonb),
                     'engine-vTEST', :uid,
                     cast('{}' as jsonb), cast('{}' as jsonb),
                     'b' || repeat('2', 63))
                """
            ),
            {
                "id": str(scen_id),
                "ws": str(workspace_id),
                "emp": str(empresa_id),
                "uid": str(user_id),
                "inp": json.dumps(
                    {"palancas": {"dep_instantanea": "10000000"}}
                ),
                "out": json.dumps(
                    {
                        "base": {"carga_total": "20000000"},
                        "simulado": {"carga_total": "17000000"},
                        "ahorro_total": "3000000",
                    }
                ),
            },
        )
        await admin_session.execute(
            text(
                """
                insert into tax_data.sii_sync_log
                    (workspace_id, empresa_id, provider, kind, status,
                     period_from, period_to, rows_inserted,
                     started_at, finished_at, created_by)
                values
                    (:ws, :emp, 'mock', 'rcv', 'success',
                     '2025-01', '2025-12', 84,
                     now() - interval '1 hour', now(), :uid)
                """
            ),
            {
                "ws": str(workspace_id),
                "emp": str(empresa_id),
                "uid": str(user_id),
            },
        )
    yield {
        "user_id": user_id,
        "workspace_id": workspace_id,
        "empresa_id": empresa_id,
        "rec_id": rec_id,
        "scen_id": scen_id,
    }
    async with admin_session.begin():
        await admin_session.execute(
            text("set local session_replication_role = 'replica'")
        )
        await admin_session.execute(
            text(
                "delete from tax_data.sii_sync_log "
                "where workspace_id = :w"
            ),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                "delete from core.recomendaciones where workspace_id = :w"
            ),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text(
                "delete from core.escenarios_simulacion "
                "where workspace_id = :w"
            ),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text("delete from core.empresas where workspace_id = :w"),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text("delete from core.workspaces where id = :w"),
            {"w": str(workspace_id)},
        )
        await admin_session.execute(
            text("delete from auth.users where id = :u"),
            {"u": str(user_id)},
        )


@pytest.mark.integration
async def test_descargar_papel_trabajo_devuelve_xlsx(
    http_client_papel: AsyncClient,
    empresa_con_diagnostico: dict[str, UUID],
) -> None:
    ctx = empresa_con_diagnostico
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(ctx["user_id"], ctx["workspace_id"])
    )
    response = await http_client_papel.get(
        f"/api/empresas/{ctx['empresa_id']}/papel-trabajo.xlsx",
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert "attachment" in response.headers.get(
        "content-disposition", ""
    )
    assert response.content[:2] == b"PK"  # XLSX = ZIP

    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    sheets = set(wb.sheetnames)
    assert {"Resumen", "Diagnostico", "Escenarios", "SII y Alertas"} <= sheets

    # Resumen incluye razón social y RUT.
    resumen = wb["Resumen"]
    flat = " ".join(
        str(cell)
        for row in resumen.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Empresa Papel SpA" in flat
    assert "11111111-1" in flat

    # Diagnóstico imprime hash de 64 chars y fundamento.
    diag = wb["Diagnostico"]
    diag_text = " ".join(
        str(cell)
        for row in diag.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "art. 14 D N°3 LIR" in diag_text
    assert "engine-vTEST" in diag_text
    # Hash plantado: 'a' + 63x'1' = 64 chars.
    expected_hash = "a" + "1" * 63
    assert expected_hash in diag_text

    # Escenarios trae el escenario plantado.
    escen = wb["Escenarios"]
    escen_text = " ".join(
        str(cell)
        for row in escen.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Plan A" in escen_text
    assert "dep_instantanea=10000000" in escen_text

    # SII y alertas.
    sii = wb["SII y Alertas"]
    sii_text = " ".join(
        str(cell)
        for row in sii.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "mock" in sii_text
    assert "84" in sii_text  # rows_inserted


@pytest.mark.integration
async def test_papel_trabajo_404_para_empresa_desconocida(
    http_client_papel: AsyncClient,
    empresa_con_diagnostico: dict[str, UUID],
) -> None:
    ctx = empresa_con_diagnostico
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(ctx["user_id"], ctx["workspace_id"])
    )
    response = await http_client_papel.get(
        f"/api/empresas/{uuid4()}/papel-trabajo.xlsx",
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 404


@pytest.mark.integration
async def test_papel_trabajo_logea_audit(
    http_client_papel: AsyncClient,
    empresa_con_diagnostico: dict[str, UUID],
    admin_session: AsyncSession,
) -> None:
    ctx = empresa_con_diagnostico
    app.dependency_overrides[verify_jwt] = _override_jwt(
        _claims(ctx["user_id"], ctx["workspace_id"])
    )
    response = await http_client_papel.get(
        f"/api/empresas/{ctx['empresa_id']}/papel-trabajo.xlsx",
        headers={"Authorization": "Bearer fake"},
    )
    assert response.status_code == 200

    async with admin_session.begin():
        result = await admin_session.execute(
            text(
                """
                select metadata
                  from security.audit_log
                 where workspace_id = :w
                   and resource_type = 'papel_trabajo'
                   and action = 'download'
                 order by at desc
                 limit 1
                """
            ),
            {"w": str(ctx["workspace_id"])},
        )
        row = result.mappings().first()
    assert row is not None
    metadata = row["metadata"]
    assert metadata["rut_masked"].startswith("11")
    assert metadata["razon_social"] == "Empresa Papel SpA"
    assert metadata["size_bytes"] > 1000
