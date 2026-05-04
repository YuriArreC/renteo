"""Renderer XLSX del papel de trabajo cliente B.

Cuatro hojas:
  1. Resumen        — identidad empresa + workspace + auditor + trace.
  2. Diagnóstico    — última recomendación con fundamento legal y hash.
  3. Escenarios     — hasta 5 simulaciones recientes con palancas + outputs.
  4. SII / Alertas  — última sincronización + alertas abiertas.

Cada hoja imprime hash + engine_version + disclaimer_version cuando
corresponde, para que el contador pueda cotejar reproducibilidad.
RUT enmascarado en logs; en el papel va completo (es el archivo del
cliente, no un canal de log).
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.papeles.builder import (
    AlertaInfo,
    EscenarioInfo,
    RecomendacionInfo,
    SiiSyncInfo,
    WorkingPaperData,
)

_HEADER_FILL = PatternFill(
    start_color="FF1F2937", end_color="FF1F2937", fill_type="solid"
)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
_MONO_FONT = Font(name="Consolas", size=9)


def _write_kv(
    ws: Worksheet, row: int, label: str, value: Any
) -> int:
    ws.cell(row=row, column=1, value=label).font = _SUBHEADER_FONT
    ws.cell(row=row, column=2, value=value)
    return row + 1


def _write_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    return row + 1


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        longest = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            longest = max(longest, len(str(value)))
        ws.column_dimensions[col_letter].width = min(
            max(longest + 2, 12), max_width
        )


def _render_resumen(
    wb: Workbook, data: WorkingPaperData
) -> None:
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Resumen")
    ws.title = "Resumen"

    ws.cell(
        row=1,
        column=1,
        value=f"Papel de trabajo — {data.empresa.razon_social}",
    ).font = Font(name="Calibri", size=14, bold=True)

    row = 3
    row = _write_kv(ws, row, "Workspace", data.workspace_name)
    row = _write_kv(ws, row, "Generado por", data.generated_by_email)
    row = _write_kv(
        ws,
        row,
        "Generado el",
        data.generated_at.strftime("%Y-%m-%d %H:%M:%S %Z").strip(),
    )
    row += 1

    row = _write_kv(ws, row, "Empresa", data.empresa.razon_social)
    row = _write_kv(ws, row, "RUT", data.empresa.rut)
    row = _write_kv(ws, row, "Giro", data.empresa.giro or "—")
    row = _write_kv(
        ws, row, "Régimen actual", data.empresa.regimen_actual
    )
    row = _write_kv(
        ws,
        row,
        "Inicio actividades",
        data.empresa.fecha_inicio_actividades or "—",
    )
    row += 1

    ws.cell(row=row, column=1, value="Disclaimer").font = _SUBHEADER_FONT
    ws.cell(
        row=row,
        column=2,
        value=(
            "Este papel de trabajo refleja un momento del motor "
            "Renteo. Las cifras son referenciales; la decisión final "
            "es responsabilidad del contador socio firmante."
        ),
    ).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    _autosize(ws)


def _render_diagnostico(
    wb: Workbook, rec: RecomendacionInfo | None
) -> None:
    ws = wb.create_sheet("Diagnostico")
    if rec is None:
        ws.cell(
            row=1,
            column=1,
            value=(
                "Sin diagnóstico de régimen registrado para esta empresa."
            ),
        )
        _autosize(ws)
        return

    row = 1
    ws.cell(row=row, column=1, value="Diagnóstico de régimen").font = (
        Font(name="Calibri", size=14, bold=True)
    )
    row += 2
    row = _write_kv(ws, row, "ID", str(rec.id))
    row = _write_kv(ws, row, "Año tributario", rec.tax_year)
    row = _write_kv(ws, row, "Tipo", rec.tipo)
    row = _write_kv(ws, row, "Régimen actual", rec.regimen_actual or "—")
    row = _write_kv(
        ws, row, "Régimen recomendado", rec.regimen_recomendado or "—"
    )
    row = _write_kv(
        ws,
        row,
        "Ahorro estimado (CLP)",
        float(rec.ahorro_estimado_clp)
        if rec.ahorro_estimado_clp is not None
        else "—",
    )
    row = _write_kv(
        ws,
        row,
        "Generada el",
        rec.created_at.strftime("%Y-%m-%d %H:%M:%S"),
    )
    row += 1

    ws.cell(row=row, column=1, value="Descripción").font = (
        _SUBHEADER_FONT
    )
    desc_cell = ws.cell(row=row, column=2, value=rec.descripcion)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    row += 2

    row = _write_header(ws, row, ["Fundamento legal"])
    if rec.fundamento_legal:
        for item in rec.fundamento_legal:
            texto = item.get("texto") or item.get("articulo") or str(item)
            ws.cell(row=row, column=1, value=str(texto)).alignment = (
                Alignment(wrap_text=True, vertical="top")
            )
            row += 1
    else:
        ws.cell(row=row, column=1, value="—")
        row += 1
    row += 1

    row = _write_header(ws, row, ["Trace de reproducibilidad"])
    cell = ws.cell(row=row, column=1, value="Hash reglas (SHA-256):")
    cell.font = _SUBHEADER_FONT
    ws.cell(row=row, column=2, value=rec.rules_snapshot_hash).font = (
        _MONO_FONT
    )
    row += 1
    row = _write_kv(ws, row, "Engine version", rec.engine_version)
    row = _write_kv(ws, row, "Disclaimer version", rec.disclaimer_version)

    _autosize(ws)


def _summarize_palancas(inputs: dict[str, Any]) -> str:
    palancas = inputs.get("palancas") or {}
    if not isinstance(palancas, dict) or not palancas:
        return "—"
    parts: list[str] = []
    for k, v in palancas.items():
        parts.append(f"{k}={v}")
    return "; ".join(parts)


def _render_escenarios(
    wb: Workbook, escenarios: list[EscenarioInfo]
) -> None:
    ws = wb.create_sheet("Escenarios")
    ws.cell(
        row=1, column=1, value="Escenarios simulados (últimos 5)"
    ).font = Font(name="Calibri", size=14, bold=True)
    row = 3
    headers = [
        "Nombre",
        "Año",
        "Régimen",
        "Carga base (CLP)",
        "Carga simulada (CLP)",
        "Ahorro (CLP)",
        "Recomendado",
        "Palancas",
        "Hash reglas",
        "Engine",
        "Generado",
    ]
    row = _write_header(ws, row, headers)
    if not escenarios:
        ws.cell(row=row, column=1, value="Sin escenarios persistidos.")
        _autosize(ws)
        return
    for sc in escenarios:
        outputs = sc.outputs
        base = outputs.get("base", {}) or {}
        simulado = outputs.get("simulado", {}) or {}
        ahorro = outputs.get("ahorro_total")
        ws.append(
            [
                sc.nombre,
                sc.tax_year,
                sc.regimen,
                float(base.get("carga_total", 0))
                if base.get("carga_total") is not None
                else None,
                float(simulado.get("carga_total", 0))
                if simulado.get("carga_total") is not None
                else None,
                float(ahorro) if ahorro is not None else None,
                "sí" if sc.es_recomendado else "no",
                _summarize_palancas(sc.inputs),
                sc.rules_snapshot_hash,
                sc.engine_version,
                sc.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    # Mark the hash column with mono font.
    hash_col = headers.index("Hash reglas") + 1
    for r in range(row, ws.max_row + 1):
        ws.cell(row=r, column=hash_col).font = _MONO_FONT
    _autosize(ws)


def _render_sii_alertas(
    wb: Workbook,
    sii: SiiSyncInfo | None,
    alertas: list[AlertaInfo],
) -> None:
    ws = wb.create_sheet("SII y Alertas")
    ws.cell(
        row=1, column=1, value="Trazabilidad SII y alertas abiertas"
    ).font = Font(name="Calibri", size=14, bold=True)

    row = 3
    ws.cell(
        row=row, column=1, value="Última sincronización SII"
    ).font = _SUBHEADER_FONT
    row += 1
    if sii is None:
        ws.cell(row=row, column=1, value="Sin sincronización registrada.")
        row += 2
    else:
        row = _write_kv(ws, row, "Proveedor", sii.provider)
        row = _write_kv(ws, row, "Tipo", sii.kind)
        row = _write_kv(ws, row, "Estado", sii.status)
        row = _write_kv(
            ws, row, "Período desde", sii.period_from or "—"
        )
        row = _write_kv(ws, row, "Período hasta", sii.period_to or "—")
        row = _write_kv(ws, row, "Filas insertadas", sii.rows_inserted)
        row = _write_kv(
            ws,
            row,
            "Inicio",
            sii.started_at.strftime("%Y-%m-%d %H:%M:%S"),
        )
        row = _write_kv(
            ws,
            row,
            "Fin",
            sii.finished_at.strftime("%Y-%m-%d %H:%M:%S")
            if sii.finished_at
            else "—",
        )
        row += 1

    ws.cell(row=row, column=1, value="Alertas abiertas").font = (
        _SUBHEADER_FONT
    )
    row += 1
    headers = [
        "Tipo",
        "Severidad",
        "Título",
        "Descripción",
        "Acción",
        "Estado",
        "Vence",
    ]
    row = _write_header(ws, row, headers)
    if not alertas:
        ws.cell(row=row, column=1, value="Sin alertas abiertas.")
    else:
        for a in alertas:
            ws.append(
                [
                    a.tipo,
                    a.severidad,
                    a.titulo,
                    a.descripcion,
                    a.accion_recomendada or "—",
                    a.estado,
                    a.fecha_limite or "—",
                ]
            )
    _autosize(ws)


def render_working_paper_xlsx(data: WorkingPaperData) -> bytes:
    """Produce el archivo XLSX completo en memoria."""
    wb = Workbook()
    _render_resumen(wb, data)
    _render_diagnostico(wb, data.recomendacion)
    _render_escenarios(wb, data.escenarios)
    _render_sii_alertas(wb, data.sii_last_sync, data.alertas)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
