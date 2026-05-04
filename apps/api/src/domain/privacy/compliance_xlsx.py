"""Renderers XLSX para RAT y DPIA (skill 5 — Ley 21.719).

Estructura RAT:
  Hoja 1 — Resumen (workspace, generado_por, fecha, count).
  Hoja 2 — Actividades (una fila por RAT con todas las columnas legales).

Estructura DPIA:
  Hoja 1 — Resumen.
  Hoja 2 — Evaluaciones (una fila por DPIA con riesgos y aprobación).

Ambos se descargan desde el panel del DPO. Pegables a un expediente
físico o anexables a un informe a la Agencia de Protección de Datos.
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(
    start_color="FF1F2937", end_color="FF1F2937", fill_type="solid"
)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True)
_SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)


def _write_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
    return row + 1


def _write_kv(ws: Worksheet, row: int, label: str, value: Any) -> int:
    ws.cell(row=row, column=1, value=label).font = _SUBHEADER_FONT
    ws.cell(row=row, column=2, value=value)
    return row + 1


def _autosize(ws: Worksheet, max_width: int = 70) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        longest = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            longest = max(longest, len(str(value).split("\n")[0]))
        ws.column_dimensions[col_letter].width = min(
            max(longest + 2, 14), max_width
        )


def _stringify_list(value: Any) -> str:
    if isinstance(value, list):
        if not value:
            return "—"
        if all(isinstance(v, str) for v in value):
            return "; ".join(value)
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return "—"
    return str(value)


def _render_resumen(
    wb: Workbook,
    *,
    title: str,
    workspace_name: str,
    generated_at: datetime,
    record_count: int,
    legal_basis: str,
) -> None:
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Resumen")
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT

    row = 3
    row = _write_kv(ws, row, "Workspace", workspace_name)
    row = _write_kv(
        ws,
        row,
        "Generado el",
        generated_at.strftime("%Y-%m-%d %H:%M:%S"),
    )
    row = _write_kv(ws, row, "Cantidad de registros", record_count)
    row = _write_kv(ws, row, "Fundamento legal", legal_basis)
    row += 1
    ws.cell(row=row, column=1, value="Disclaimer").font = _SUBHEADER_FONT
    cell = ws.cell(
        row=row,
        column=2,
        value=(
            "Documento generado por Renteo para apoyar el cumplimiento "
            "de la Ley 21.719. La revisión y firma final son "
            "responsabilidad del DPO designado del workspace."
        ),
    )
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    _autosize(ws)


def render_rat_xlsx(
    *,
    records: list[dict[str, Any]],
    workspace_name: str,
    generated_at: datetime,
) -> bytes:
    wb = Workbook()
    _render_resumen(
        wb,
        title="Registro de Actividades de Tratamiento (RAT)",
        workspace_name=workspace_name,
        generated_at=generated_at,
        record_count=len(records),
        legal_basis="art. 15-16 Ley 21.719",
    )

    ws = wb.create_sheet("Actividades")
    headers = [
        "ID",
        "Nombre actividad",
        "Finalidad",
        "Base legal",
        "Categorías titulares",
        "Categorías datos",
        "Datos sensibles",
        "Encargados referenciados",
        "Transferencias internacionales",
        "Plazo conservación",
        "Medidas seguridad",
        "Responsable (email)",
        "Creado",
        "Actualizado",
        "Archivado",
    ]
    row = _write_header(ws, 1, headers)
    if not records:
        ws.cell(row=row, column=1, value="Sin actividades registradas.")
        _autosize(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    for rec in records:
        ws.append(
            [
                str(rec["id"]),
                rec.get("nombre_actividad", ""),
                rec.get("finalidad", ""),
                rec.get("base_legal", ""),
                _stringify_list(rec.get("categorias_titulares")),
                _stringify_list(rec.get("categorias_datos")),
                "sí" if rec.get("datos_sensibles") else "no",
                _stringify_list(rec.get("encargados_referenciados")),
                _stringify_list(
                    rec.get("transferencias_internacionales")
                ),
                rec.get("plazo_conservacion", ""),
                _stringify_list(rec.get("medidas_seguridad")),
                rec.get("responsable_email", ""),
                rec.get("created_at", ""),
                rec.get("updated_at", ""),
                rec.get("archived_at") or "—",
            ]
        )
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_dpia_xlsx(
    *,
    records: list[dict[str, Any]],
    workspace_name: str,
    generated_at: datetime,
) -> bytes:
    wb = Workbook()
    _render_resumen(
        wb,
        title="Evaluaciones de Impacto en Protección de Datos (DPIA)",
        workspace_name=workspace_name,
        generated_at=generated_at,
        record_count=len(records),
        legal_basis="art. 35 Ley 21.719",
    )

    ws = wb.create_sheet("Evaluaciones")
    headers = [
        "ID",
        "RAT vinculado",
        "Nombre evaluación",
        "Descripción tratamiento",
        "Necesidad / proporcionalidad",
        "Riesgos identificados",
        "Medidas mitigación",
        "Riesgo residual",
        "Aprobado por DPO",
        "Aprobado el",
        "Versión",
        "Creado",
        "Actualizado",
    ]
    row = _write_header(ws, 1, headers)
    if not records:
        ws.cell(row=row, column=1, value="Sin evaluaciones registradas.")
        _autosize(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    for rec in records:
        ws.append(
            [
                str(rec["id"]),
                str(rec.get("rat_id") or "—"),
                rec.get("nombre_evaluacion", ""),
                rec.get("descripcion_tratamiento", ""),
                rec.get("necesidad_proporcionalidad", ""),
                _stringify_list(rec.get("riesgos_identificados")),
                _stringify_list(rec.get("medidas_mitigacion")),
                rec.get("riesgo_residual", ""),
                rec.get("aprobado_por_dpo_email") or "—",
                rec.get("aprobado_at") or "—",
                rec.get("version", 1),
                rec.get("created_at", ""),
                rec.get("updated_at", ""),
            ]
        )
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
